"""Extract content from Excel and Word files"""
import openpyxl
from docx import Document

# File paths
base_path = r"C:\Users\serge\Desktop\retirement-simulator-dev\retirement-simulator\DOCUMENTS RETIREMENT APP"
files = {
    'excel': f"{base_path}\\app_py future dev features.xlsx",
    'docx1': f"{base_path}\\FUTURE Additional features GROK.docx",
    'docx2': f"{base_path}\\GROK_Human-Level Report for Project Planning.docx",
    'docx3': f"{base_path}\\GROK_Tech Report for AI Model Instr.docx"
}

output_path = r"C:\Users\serge\Desktop\retirement-simulator-dev\retirement-simulator\DOCUMENTS RETIREMENT APP"

# Extract Excel
print("Extracting Excel file...")
wb = openpyxl.load_workbook(files['excel'])
with open(f"{output_path}\\extracted_excel.txt", 'w', encoding='utf-8') as f:
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        f.write(f"=== SHEET: {sheet_name} ===\n\n")
        for row in sheet.iter_rows(values_only=True):
            f.write('\t'.join([str(cell) if cell is not None else '' for cell in row]) + '\n')
        f.write('\n')
print("Excel extracted!")

# Extract Word documents
for i, (key, filepath) in enumerate([('docx1', files['docx1']), ('docx2', files['docx2']), ('docx3', files['docx3'])], 1):
    print(f"Extracting Word document {i}/3...")
    doc = Document(filepath)
    output_file = f"{output_path}\\extracted_word_{i}.txt"
    with open(output_file, 'w', encoding='utf-8') as f:
        for para in doc.paragraphs:
            f.write(para.text + '\n')
    print(f"Word document {i} extracted!")

print("\nAll files extracted successfully!")
print(f"Output location: {output_path}")
